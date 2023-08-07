# Task 1
import openpyxl
from openpyxl import workbook

def generatePreferences(values): 
  """The function extracts values from the worksheet 
    from the openpyxl library. """
  preferences = {} 
  """creates an empty dictionary called preferences 
  and initializes two variables, agentValues and alternativeValues, 
  which represent the number of rows and columns in the values sheet, as seen below."""
  
  agentValues = values.max_row #variable for number of rows
  alternativeValues = values.max_column #variable for number of columns
  
#Loop iterating through the range of rows on the value worksheet
  for i in range(1,agentValues+1): 
    preferenceOrdering = {}
    
    #Loop iterating through the range of columns on the value worksheet
    for j in range(1, alternativeValues+1): 
      
      valuations = values.cell(row=i, column=j).value
      """valuation refers to the numerical value of the cell i,j that 
      represents how happy an agent would be if that specific alternative 
      were to be selected"""

      preferenceOrdering[j] = valuations #order of preference for alternatives
      preferences[i] = preferenceOrdering

  # this sorts out the preferences of each agent by value in descending order
  for k, value in preferences.items():
    preferences[k] = [i[0] for i in sorted(value.items(),key = lambda x:(x[1],x[0]),reverse =True)]
  
  return preferences
"""returns the preferences containing the lists of alternatives in the worsheet
   in order of preference for each Agent."""

# Task 2
def dictatorship(preferenceProfile, agent):
    """This function takes into account the preferenceProfile and Agent
    by first checking if the agent is present in the preferenceProfile .
    
    If the agent is present, the function returns the first item 
    in the list of preferences, which would be considered the winner,
    
    if it is not,then it raises value error."""
    if agent not in preferenceProfile:
        raise ValueError("The preference profile does not correspond to any agent")

    return preferenceProfile[agent][0]


def scoringRule (preferences, scoreVector, tieBreak) -> int:
    """This function takes into account the preferences, scoreVector, and tieBreak,
    by first determineing the number of agents and alternatives in the preferences.
     """
    numAgents = len(preferences.keys())
    numAlternatives = len(preferences[1])
    try:
    #checks if the length of the scoreVector list is equal to the number of alternatives
        if len(scoreVector) != numAlternatives:
            #raises a ValueError if it is not equal
            raise ValueError('Incorrect input')
    except ValueError as e:
        # Prints error message and return False 
        print(e)
        return False
    else:
        #sorts the scoreVector in reverse order
        scoreVector = sorted(scoreVector, reverse=True) 
        #creates a list of preference lists from the preferences dictionary
        prefList = list(preferences.values())
        """creates a dictionary that will hold the total scores 
        for each alternative"""
        scores = {}
        for pref in prefList: 
            for i,j in enumerate(pref):
                """adds the corresponding score from the scoreVector 
                to the scores dictionary for each alternative."""
                scores[j] = scores.get(j, 0) + scoreVector[i]
        all_values = scores.values()
        max_score = max(all_values) #finds maximum scores in scores dictionary
        #creates a list of alternatives with that maximum score called winners
        winners = [a for a, current_score in scores.items() if current_score == max_score]
    #calls the tieBreaker function to resolve the tie when there is more than one winner
    if len(winners) > 1:
        try:
            return tieBreaker(tieBreak, winners, preferences)
        except Exception as e:
            print("Please try again, an error has occurred the tieBreaker function:", e)
        return None
    return winners[0] #return fuction if only one winner or tie is resolved


def plurality(preferences, tieBreak) -> int:
    """This function takes into account preferences and tieBreak """
    scores = {} #hold the number of votes for each alternative
    for agent, pref in preferences.items():
        if pref[0] in scores:
            #adds 1 to the scores dictionary for the first preference of each agent
            scores[pref[0]] += 1
        else:
            scores[pref[0]] = 1
    #finds the maximum number of votes in the scores dictionary
    max_score = max(scores.values())
    #creates a list -Winners- containing the alternatives with the maximum number of votes
    Winners = [a for a, current_score in scores.items() if current_score == max_score]
    #if more than one winner, the function calls the tieBreaker function with the tieBreak
    if len(Winners) > 1: 
        try:
            return tieBreaker(tieBreak, Winners, preferences)
            #If an error occurs when calling the tieBreaker function, it prints the error message
        except Exception as e:
            print("An error occurred when calling the tieBreaker function:", e)
        return None
    return Winners[0]


def veto(preferences, tieBreak) -> int:
    """This function takes into account preference and tieBreak"""
    
    #  stores the scores for each alternative
    scores = {}
     # Iterate through each alternative in the preference list
    for agent, pref in preferences.items():
        for i in range(len(pref)):
            #retrives 
            preference = pref[i]
            if preference not in scores:
                scores[preference] = 0
            if i < len(pref) - 1:
                """adds 1 to the scores dictionary for each alternative 
                that is ranked higher than another alternative"""
                scores[preference] += 1
    if not scores:
        return None
    max_score = max(scores.values())
    #If the maximum number of vetoes is 0, the function also returns None
    if max_score == 0:
        return None
    winners = [a for a, current_score in scores.items() if current_score == max_score]
    
    #if more than one winner, the function calls the tieBreaker function with the tieBreak
    if len(winners) > 1:
        try:
            return tieBreaker(tieBreak, winners, preferences)
        except Exception as e:
            print("An error occurred when calling the tieBreaker function:", e)
        #If an error occurs when calling the tieBreaker function, it prints the error message and returns None
        return None
    return winners[0]


def borda(preferences, tieBreak) -> int:
    scores = {}
    for agent, pref in preferences.items():
        for i in range(len(pref)):
            preference = pref[i]
            if preference not in scores:
                scores[preference] = 0
            scores[preference] += i
    if not scores:
        return None
    min_score = min(scores.values())
    if min_score == 0:
        return None
    winners = [a for a, current_score in scores.items() if current_score == min_score]
    ## If there is more than one winner, use the tie-breaking rule to select a single winner

    if len(winners) > 1:
        try:
            return tieBreaker(tieBreak, winners, preferences)
        except Exception as e:
                # Print an error message if an exception is raised while calling the tieBreaker function
            print("An error occurred when calling the tieBreaker function:", e)
        return None
    return winners[0]


def harmonic(preferences, tieBreak) -> int:
    """"initializing an empty dictionary scores, where the keys are the candidates and the values are the scores for each candidate. It then iterates over the keys and values in 
    preferences and assigns a score to each candidate based on their rank in the preference list."""
    scores = {}
    for agent, pref in preferences.items():
        for i in range(1, len(pref) + 1):
            preference = pref[i - 1]
            if preference not in scores:
                scores[preference] = 0
            scores[preference] += 1 / i
    max_score = max(scores.values())
    winners = [a for a, current_score in scores.items() if current_score == max_score]
    """If there is more than one candidate with the maximum score, the function calls the tieBreaker function with tieBreak, a list of the candidates with the maximum score, and preferences as arguments, and returns the result. 
    If the call to tieBreaker raises an exception, the function prints an error message and returns None"""
    if len(winners) > 1:
        try:
            return tieBreaker(tieBreak, winners, preferences)
        except Exception as e:
            print("An error occurred when calling the tieBreaker function:", e)
        return None
    return winners[0]


import copy 

def STV(preferences, tieBreak) -> int:
    #
    preference_list = copy.deepcopy(list(preferences.values()))
    deleted_values = []
    votes_counts = [0] * (len(preferences[1]) + 1)
    size = len(preference_list[0])
    """"enters a loop that continues until size is 0. Inside the loop, the function counts the votes for each candidate by adding 1 to the corresponding element in votes_counts for each appearance of the candidate in preference_list. 
    It then creates a dictionary scores where the keys are the number of votes received by each candidate and the values are lists of the candidates that received that number of votes."""
    while (size > 0):
        for row in preference_list:
            votes_counts[row[0]] += 1
        scores = {}
        tied_candidates = []
        for index in range(len(votes_counts)):
            if votes_counts[index] in scores.keys():
                tied_candidates[:] = scores[votes_counts[index]]
            if (index != 0):
                tied_candidates.append(index)
                scores[votes_counts[index]] = tied_candidates[:]
                tied_candidates = []
        minimum_key = min(scores.keys())
        """If there are two or fewer keys in scores, the function calls the tieBreaker function with tieBreak, the list of candidates that received the minimum number of votes, and preferences as arguments, and returns the result. If there are more than two keys in scores, the function removes the candidates that received the minimum number of votes from preference_list and votes_counts, and appends them to deleted_values. 
        It then resets all the elements in votes_counts to 0 if they were infinity, and continues the loop."""
        if (len(scores.keys()) <= 2):
            winning_value = tieBreaker(tieBreak, scores[minimum_key], preferences)
            return winning_value
        else:
            for val in scores[minimum_key]:
                if (val not in deleted_values):
                    for i, values in enumerate(preference_list):
                        values.remove(val)
                        deleted_values.append(val)
                        votes_counts[val] = float('inf')
        votes_counts = [float('inf') if row == float('inf') else 0 for row in votes_counts]
    return 0

def rangeVoting(values, tieBreak) -> int:
    """generates a dictionary where the keys are the agents 
    and the values are lists of the alternatives sorted in order of the agent's preference."""
    agentDictionary = generatePreferences(values)
    alternativeScores = {}
    #a list of the scores for each alternative all initialized to 0
    votesCounts = [0] * (len(list(agentDictionary.values())[0]) + 1)
    
    #iterates over the values in the dictionary of preferences and their corresponding values in values
    for i, value in enumerate(agentDictionary.values()):
        numerical_values = [row.value for row in list(values.rows)[i]]
        #counts the votes for each alternative
        for v in value:
            votesCounts[v] += float(numerical_values[v - 1])
    count_list = []
    # Iterate the votes counts for each alternative
    for index in range(len(votesCounts)):
        """If an alternative has received more votes than any other alternative, the function calls the 
        tieBreaker function with tieBreak, the list of alternatives that received the maximum number of votes, and the dictionary of preferences as arguments, and returns the result. If the call to tieBreaker raises an exception, the function prints an error message and returns None."""

        if votesCounts[index] in alternativeScores.keys():
            count_list[:] = alternativeScores[votesCounts[index]]

        if (votesCounts[index] != 0):
            count_list.append(index)
            alternativeScores[votesCounts[index]] = count_list[:]
            count_list = []

    max_occurence = max(alternativeScores.keys())
    try:
        return tieBreaker(tieBreak, alternativeScores[max_occurence], agentDictionary)
    except:
        print("Invalid tie-breaking rule")
        return None

def tieBreaker(tiebreak, winners, preferences) -> int:
    if tiebreak == "max": #returns the maximum value in the winners list.
        return max(winners)
    elif tiebreak == "min": #returns the minimum value in the winners list.
        return min(winners)
    else:
        #the function tries to convert it to an integer and store the result in agentValidity
        try:
            agentValidity = int(tiebreak)
            #f this conversion is successful, the function then checks that agentValidity is a valid key in the preferences dictionary.
            assert agentValidity in preferences.keys()
            
            maximumRank = float('inf')
            maximumNumber = -1
            """the function looks for the element in winners that appears first in the list stored 
            as the value for agentValidity in preferences, and returns that element."""

            for row in winners:
                if (preferences[agentValidity].index(row) < maximumRank):
                    maximumRank = preferences[agentValidity].index(row)
                    maximumNumber = row
            return maximumNumber
            #if it does not return anything then it raises a ValueError exception.
        except AssertionError:
            print(f"{tiebreak} does not correspond to an agent.")
        except ValueError:
            print(f"{tiebreak} could not be converted to an integer.")